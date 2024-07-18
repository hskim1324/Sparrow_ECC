import openpyxl, time
import torch

## analyze model weights

device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')

# toggle tensorfloat32 in Ampere or later GPUs
if torch.cuda.is_available():
    torch.backends.cuda.matmul.allow_tf32 = False

####################################
PATH = '' # TODO: fill in path to model weight file
####################################

dict_params = torch.load(PATH, map_location=device)

diff_count = [0 for i in range(129)]

block_size = 64 # weight block size to analyze: 64 weights
remain = 0

remainder = torch.empty(1, device=device)

with torch.inference_mode():
    for name, _ in dict_params.items():
        if dict_params[name].dtype != torch.bfloat16:
            dict_params[name] = dict_params[name]
            continue

        if 'weight' not in name:
            if 'bias' not in name:
                dict_params[name] = dict_params[name]
                continue

        temp = dict_params[name]
        temp = temp.to(torch.float32)
        temp = temp.view(torch.int32)
        temp = temp.view(dict_params[name].numel())
        
        temp[temp==0] = 0x7F800000

        exponents = (temp & 0x7F800000) >> 23

        del temp
        torch.cuda.empty_cache()

        if remain == 1:
            exponents_old = exponents
            exponents = torch.cat((remainder, exponents))
            del exponents_old
            torch.cuda.empty_cache()
            remain = 0

        rows = int(exponents.numel() / block_size)
        
        if exponents.numel() % block_size != 0:
            remain = 1
            remainder = exponents[rows*block_size:].clone().detach()
        
        exponents = exponents.view(rows, block_size)

        # exponents[exponents==0xFF] = 0 # use to exclude 0 values, default: include 0 values

        max_expos, _ = torch.max(exponents, -1, True)

        # exponents[exponents==0] = 0xFF # use to exclude 0 values, default: include 0 values

        min_expos, _ = torch.min(exponents, -1, True)

        diff_expos = max_expos - min_expos
        diff_expos[diff_expos < 0] = 0

        diff_expos = diff_expos.view(diff_expos.numel())

        for i in range(129):
            diff_count[i] += int(diff_expos[diff_expos == i].numel())
        
        del max_expos, min_expos, diff_expos
        torch.cuda.empty_cache()
    
        print("done for layer: " + name)

total_count = 0
for i in range(129):
    total_count += diff_count[i]

diff_proportion = [0 for i in range(129)]
for i in range(129):
    diff_proportion[i] = (int(diff_count[i]) / total_count) * 100

## generate result data excel file

exponent_diff_excel_filename = 'exponent_diff_' + time.strftime('%y%m') + '.xlsx'
wb = openpyxl.Workbook()

wb.active.title = 'Count'
wb.create_sheet('Proportion')
meta_names = ['Model']

for i in range(129):
    meta_names.append(str(i))

wb['Count'].append(meta_names)
wb['Proportion'].append(meta_names)
wb.save(exponent_diff_excel_filename)
wb.close()

## write result data

wb = openpyxl.load_workbook(exponent_diff_excel_filename)
write_row_1 = ['llama3']
write_row_2 = ['llama3']
for i in range(129):
    write_row_1.append(diff_count[i])
    write_row_2.append(diff_proportion[i])
wb['Count'].append(write_row_1)
wb['Proportion'].append(write_row_2)
wb.save(exponent_diff_excel_filename)
wb.close()

torch.cuda.empty_cache()